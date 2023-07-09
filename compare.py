import sys
import tarfile
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
from datetime import datetime, timedelta

dates = []
today = datetime.strptime(sys.argv[1], '%Y%m%d')
ndays = int(sys.argv[2])
for i in range(ndays):
    dates.append(datetime.strftime(today, '%Y%m%d'))
    today = today - timedelta(1)

def change_row_font(min_row, max_row, min_col, max_col, font):
    for new_rows in ws.iter_rows(min_row = min_row, max_row = max_row, min_col = min_col, max_col = max_col):
        for new_cell in new_rows:
            new_cell.font = font

wb = Workbook()
ws = wb.active

deleted_font = Font(color = 'FF0000')
added_font = Font(color = '339966')

prefix = '/home/junlin/products/kramp'
i = 0
while i < (ndays - 1):
    ws.title = dates[i]

    filename_a = "{0}/kramp_{1}.csv.tar.bz2".format(prefix, dates[i])
    filename_b = "{0}/kramp_{1}.csv.tar.bz2".format(prefix, dates[i + 1])
    lines_a = set()
    lines_b = set()
    print("comparing {0} and {1}".format(filename_a, filename_b))

    with tarfile.open(filename_a, 'r:bz2') as tar_a:
        f = tar_a.extractfile(tar_a.members[0])
        for l in f.readlines():
            lines_a.add(l.decode('utf8'))
        f.close()

    with tarfile.open(filename_b, 'r:bz2') as tar_b:
        f = tar_b.extractfile(tar_b.members[0])
        for l in f.readlines():
            l_utf8 = l.decode('utf8')
            try:
                lines_a.remove(l_utf8)
            except:
                lines_b.add(l_utf8)
        f.close()
        
    current_row = 1
    for l in lines_a:
        cells = l.split(';')
        ws.append(cells)
        change_row_font(current_row, current_row, 1, len(cells), deleted_font)

        current_row += 1

    for l in lines_b:
        cells = l.split(';')
        ws.append(cells)
        change_row_font(current_row, current_row, 1, len(cells), added_font)

        current_row += 1

    i += 1

    if i < (ndays - 1):
        ws = wb.create_sheet()

wb.save('report.xlsx')
